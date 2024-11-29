"""
Student Exam Portal Application
------------------------------
A Streamlit application for managing online exams with features including:
- Student authentication
- Question paper distribution
- Timed exam sessions
- Answer sheet submission
- Solution access
- Activity logging
"""

import streamlit as st
import time
from streamlit_autorefresh import st_autorefresh
from google.oauth2 import service_account
from google.cloud import storage
import mimetypes
import io
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import logging
import traceback

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize session state variables
def init_session_state():
    """Initialize all session state variables with defaults."""
    if 'timer_running' not in st.session_state:
        st.session_state.timer_running = False
    if 'start_time' not in st.session_state:
        st.session_state.start_time = None
    if 'elapsed_time' not in st.session_state:
        st.session_state.elapsed_time = 0
    if 'exam_duration' not in st.session_state:
        st.session_state.exam_duration = 6000  # 100 minutes in seconds
    if 'answer_uploaded' not in st.session_state:
        st.session_state.answer_uploaded = False
    if 'question_downloaded' not in st.session_state:
        st.session_state.question_downloaded = False
    if 'solution_downloaded' not in st.session_state:
        st.session_state.solution_downloaded = False
    if 'time_up' not in st.session_state:
        st.session_state.time_up = False
    if 'download_question_clicked' not in st.session_state:
        st.session_state.download_question_clicked = False
    if 'upload_answer_clicked' not in st.session_state:
        st.session_state.upload_answer_clicked = False
    if 'download_solution_clicked' not in st.session_state:
        st.session_state.download_solution_clicked = False
    if "page" not in st.session_state:
        st.session_state["page"] = "login"
    if "login_time" not in st.session_state:
        st.session_state.login_time = None
    if "upload_progress" not in st.session_state:
        st.session_state.upload_progress = 0
    if 'is_transitioning' not in st.session_state:
        st.session_state.is_transitioning = False
    if 'loading_message' not in st.session_state:
        st.session_state.loading_message = ""

def change_page(new_page, loading_message=""):
    """Handle smooth page transitions."""
    st.session_state.is_transitioning = True
    st.session_state.loading_message = loading_message
    st.session_state.page = new_page

# GCS Operations with retry mechanism
@st.cache_resource
def get_gcs_client():
    """Initialize Google Cloud Storage client with credentials from Streamlit secrets."""
    try:
        credentials = service_account.Credentials.from_service_account_info(
            st.secrets["gcp_service_account"]
        )
        return storage.Client(credentials=credentials)
    except Exception as e:
        logger.error(f"Failed to initialize GCS client: {str(e)}")
        return None

@st.cache_data(ttl=300)
def cached_download_from_gcs(folder_path, filename):
    """Download a file from Google Cloud Storage with caching."""
    try:
        client = get_gcs_client()
        if not client:
            return None
            
        bucket = client.bucket(st.secrets["gcs"]["bucket_name"])
        blob = bucket.blob(f"{folder_path}/{filename}")
        
        if not blob.exists():
            return None
            
        return blob.download_as_bytes()
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return None

def upload_to_gcs(file_data, file_name, folder_path):
    """Upload file to GCS with retry mechanism."""
    try:
        client = get_gcs_client()
        if not client:
            return False
            
        bucket = client.bucket(st.secrets["gcs"]["bucket_name"])
        blob = bucket.blob(f"{folder_path}/{file_name}")
        
        file_data.seek(0)
        content_type = mimetypes.guess_type(file_name)[0]
        if content_type:
            blob.content_type = content_type
        
        blob.upload_from_file(file_data)
        return True
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return False

def log_student_activity(name, email, class_name, test_number, is_login=True):
    """Log student activity (login/logout) to Excel."""
    try:
        log_content = cached_download_from_gcs("logs", "student_logs.xlsx")
        if log_content is None:
            return False
            
        wb = load_workbook(BytesIO(log_content))
        ws = wb.active
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if is_login:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=next_row-1)
            ws.cell(row=next_row, column=2, value=name)
            ws.cell(row=next_row, column=3, value=email)
            ws.cell(row=next_row, column=4, value=class_name)
            ws.cell(row=next_row, column=5, value=test_number)
            ws.cell(row=next_row, column=6, value=current_time)
            st.session_state.login_time = current_time
        else:
            login_time = st.session_state.login_time
            for row in range(2, ws.max_row + 1):
                if (ws.cell(row=row, column=2).value == name and 
                    ws.cell(row=row, column=3).value == email and
                    ws.cell(row=row, column=4).value == class_name and
                    ws.cell(row=row, column=5).value == test_number and
                    ws.cell(row=row, column=6).value == login_time):
                    ws.cell(row=row, column=7, value=current_time)
                    login_dt = datetime.strptime(login_time, "%Y-%m-%d %H:%M:%S")
                    logout_dt = datetime.strptime(current_time, "%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row, column=8, value=str(logout_dt - login_dt))
                    break
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return upload_to_gcs(excel_bytes, "student_logs.xlsx", "logs")
    except Exception as e:
        logger.error(f"Logging error: {str(e)}")
        return False

def handle_login_submit():
    """Handle student login form submission."""
    name = st.session_state.get("student_name", "").strip()
    email = st.session_state.get("student_email", "").strip()
    
    if not name:
        st.error("Please enter your name")
        return
    if not email:
        st.error("Please enter your email ID")
        return
        
    st.session_state["name"] = name
    st.session_state["email"] = email
    st.session_state["class"] = f"Grade{st.session_state.student_class}"
    st.session_state["test"] = st.session_state.test_number
    
    with st.spinner('Logging in...'):
        if log_student_activity(name, email, st.session_state["class"], st.session_state["test"]):
            change_page("download_questions", "Loading questions...")
        else:
            st.error("Failed to log student information. Please try again.")

def handle_question_download():
    """Handle question paper download."""
    st.session_state.download_question_clicked = True
    st.session_state.question_downloaded = True
    change_page("upload_answer_sheet", "Preparing exam environment...")

def stop_timer():
    """Stop the exam timer."""
    if st.session_state.timer_running:
        st.session_state.timer_running = False
        st.session_state.elapsed_time = time.time() - st.session_state.start_time
        if st.session_state.elapsed_time >= st.session_state.exam_duration:
            st.session_state.time_up = True

def main():
    """Main application logic."""
    init_session_state()
    
    # Show loading state during transitions
    if st.session_state.is_transitioning:
        with st.spinner(st.session_state.loading_message):
            st.session_state.is_transitioning = False
            time.sleep(0.5)  # Brief pause for smooth transition
            st.rerun()
    
    # Login Page
    if st.session_state.page == "login":
        st.title("Student Login")
        with st.form("login_form", clear_on_submit=False):
            st.text_input("Student Name", key="student_name")
            st.text_input("Email ID", key="student_email")
            st.selectbox("Class", [9, 10, 11, 12], key="student_class")
            st.selectbox("Test Number", ["Test1", "Test2", "Test3"], key="test_number")
            st.form_submit_button("Submit", on_click=handle_login_submit)
    
    # Download Questions Page
    elif st.session_state.page == "download_questions":
        st.title("Download Questions")
        st.subheader("Download Question Paper")
        
        question_folder = f"{st.session_state['class']}/{st.session_state['test']}/QuestionPaper"
        question_content = cached_download_from_gcs(question_folder, "QuestionPaper.pdf")
        
        if question_content is not None:
            if not st.session_state.download_question_clicked:
                st.download_button(
                    label="DOWNLOAD QUESTION PAPER",
                    data=question_content,
                    file_name="QuestionPaper.pdf",
                    mime="application/pdf",
                    on_click=handle_question_download
                )
        else:
            st.error("Question paper not available in Google Cloud Storage. Please contact your administrator.")
    
    # Upload Answer Sheet Page
    elif st.session_state.page == "upload_answer_sheet":
        st.title("Upload Answer Sheet")
        
        if not st.session_state.timer_running:
            st.session_state.timer_running = True
            st.session_state.start_time = time.time()
        
        if st.session_state.timer_running:
            st.session_state.elapsed_time = time.time() - st.session_state.start_time
        
        remaining_time = max(st.session_state.exam_duration - st.session_state.elapsed_time, 0)
        if remaining_time <= 0:
            st.session_state.time_up = True
            st.session_state.timer_running = False
        
        minutes = int(remaining_time // 60)
        seconds = int(remaining_time % 60)
        time_str = f"{minutes:02d}:{seconds:02d}"
        st.subheader(f"Time Remaining: ⏱️ {time_str}")
        
        upload_disabled = st.session_state.time_up or st.session_state.upload_answer_clicked
        
        uploaded_file = st.file_uploader(
            "Choose a PDF file to upload your answer sheet",
            type="pdf",
            key="answer_upload",
            disabled=upload_disabled
        )
        
        if st.button("Submit Answer Sheet", disabled=upload_disabled):
            if uploaded_file:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    status_text.text('Preparing submission...')
                    stop_timer()
                    progress_bar.progress(20)
                    
                    status_text.text('Uploading your answer sheet...')
                    file_path = f"{st.session_state['class']}/{st.session_state['test']}/AnswerSheets"
                    filename = f"{st.session_state['name'].replace(' ', '_')}_{st.session_state['class']}.pdf"
                    
                    if upload_to_gcs(uploaded_file, filename, file_path):
                        progress_bar.progress(60)
                        status_text.text('Updating records...')
                        
                        if log_student_activity(
                            st.session_state["name"], 
                            st.session_state["email"],
                            st.session_state["class"], 
                            st.session_state["test"],
                            False
                        ):
                            progress_bar.progress(90)
                            status_text.text('Finalizing...')
                            st.session_state.answer_uploaded = True
                            st.session_state.upload_answer_clicked = True
                            progress_bar.progress(100)
                            status_text.text('Complete!')
                            st.success("Answer sheet uploaded successfully!")
                            time.sleep(1)
                            change_page("thank_you")
                            st.rerun()
                        else:
                            st.error("Failed to update records. Please try again.")
                    else:
                        st.error("Failed to upload answer sheet. Please try again.")
                except Exception as e:
                    logger.error(f"Error during submission process: {str(e)}")
                    st.error("An error occurred during submission. Please try again.")
            else:
                st.error("Please select a file to upload")
        
        if st.session_state.timer_running and not st.session_state.answer_uploaded:
            st_autorefresh(interval=2000, key="timer_refresh")
    
    # Thank You Page
    elif st.session_state.page == "thank_you":
        st.title("THANK YOU")
        st.header(f"Thank you {st.session_state['name']} for completing the test!")
        st.success("Your answer sheet has been submitted successfully.")
        
        solutions_folder = f"{st.session_state['class']}/{st.session_state['test']}/Solutions"
        solution_content = cached_download_from_gcs(solutions_folder, "Solutions.pdf")
        
        if solution_content is not None and not st.session_state.download_solution_clicked:
            if st.download_button(
                label="DOWNLOAD SOLUTIONS",
                data=solution_content,
                file_name="Solutions.pdf",
                mime="application/pdf"
            ):
                st.session_state.solution_downloaded = True
                st.session_state.download_solution_clicked = True
                st.success("Solutions downloaded successfully!")
        else:
            st.error("Solutions not available in Google Cloud Storage. Please contact your administrator.")

if __name__ == "__main__":
    main()